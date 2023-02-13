import tkinter as tk
from tkinter import filedialog as fd
from datetime import datetime as dt, datetime
import openpyxl as pyxl
from openpyxl.styles import Alignment
import shutil
import re


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.button_configs = [('Год, за который оформляется КУД', self.change_year, 0),
                               ('Исходный xlsx файл в котором оформляется КУД)', self.pattern, 1),
                               ('xlsx файл для загрузки данных ОФД', self.in_file, 2)]
        self.opts = {'padx': 10, 'pady': 10, 'ipadx': 10, 'ipady': 10, 'sticky': 'nswe'}
        self.geometry('905x480+200+100')
        self.title("Книга учета доходов (КУД) ИП, применяющих патентную систему налогообложения")
        self.btns = [self.create_btn(text, command, row) for text, command, row in self.button_configs]
        self.year = tk.StringVar()
        self.year.trace("w", self.show_year)
        self.pattern0 = tk.StringVar()
        self.pattern0.set("Исходный xlsx файл в котором оформляется КУД")
        self.year_ent = tk.Entry(self, textvariable=self.year, justify='center')
        self.year.set("2000")
        self.save_year_btn = tk.Button(text='Сохранить год', command=self.save_year)
        self.result_btn = tk.Button(text='Сформировать КУД с текущими параметрами', command=self.result)
        self.pattern_lbl = tk.Label(textvariable=self.pattern0, wraplength=451)
        self.in_file0 = tk.StringVar()
        self.in_file_lbl = tk.Label(textvariable=self.in_file0, wraplength=451)
        self.in_file0.set("xlsx файл для загрузки данных ОФД")
        self.result0 = tk.StringVar()
        self.result0.set("- Расположение итогового xlsx файла КУД под именем: КУД year_год-месяц-дата-время.xlsx")
        self.result_file_lbl = tk.Label(textvariable=self.result0, wraplength=700)

        self.year_ent.grid(row=0, column=1, **self.opts)
        self.save_year_btn.grid(row=0, column=2, **self.opts)
        self.pattern_lbl.grid(row=1, column=1, columnspan=2, **self.opts)
        self.in_file_lbl.grid(row=2, column=1, columnspan=2, **self.opts)
        self.result_file_lbl.grid(row=3, column=0, columnspan=3, **self.opts)
        self.result_btn.grid(row=4, columnspan=3, **self.opts)

    def create_btn(self, text, command, row):
        return tk.Button(self, text=text, command=command).grid(row=row, column=0, **self.opts)

    def change_year(self):
        self.year_ent.config(bg='white')
        self.year_ent.icursor(tk.END)
        self.year_ent.focus()

    def show_year(self, *args):
        self.year_ent.config(textvariable=self.year)

    def save_year(self):
        self.year_ent.config(bg='green')
        self.focus()

    def pattern(self):
        filetypes = (('Excel файлы', '*.xlsx'), ('Любой', '*'))
        filename = fd.askopenfilename(title='Исходный xlsx файл в котором оформляется КУД',
                                      initialdir='PATTERN', filetypes=filetypes)
        self.pattern0.set(filename)
        self.pattern_lbl.config(bg='green')

    def in_file(self):
        filetypes = (('Excel файлы', '*.xlsx'), ('Любой', '*'))
        filename = fd.askopenfilename(title='xlsx файл для загрузки данных ОФД',
                                      initialdir='DOWNLOADS', filetypes=filetypes)
        self.in_file0.set(filename)
        self.in_file_lbl.config(bg='green')

    # noinspection PyGlobalUndefined
    def result(self):
        global header_row, date_time_col, sum_col, nomer_fd_col, last_row
        result = f'RESULT/КУД{self.year_ent.get()}_{dt.now().year}-{dt.now().month}-' \
                 f'{dt.now().day}_{dt.now().hour}:{dt.now().minute}:{dt.now().second}.xlsx'
        self.result0.set(f'Расположение итогового xlsx файла КУД: {result}')
        # Обновляем значение result0 для result_file_lbl
        self.result_file_lbl.config(bg='green')
        shutil.copyfile(self.pattern0.get(), result)

# Получение названий листов из файла Excel
        result_file = pyxl.load_workbook(result)
        ofd_file = pyxl.load_workbook(self.in_file0.get())

# Получение специфического листа из файла Excel
        result_data_sheet = result_file[result_file.sheetnames[1]]
        data_sheet = ofd_file[ofd_file.sheetnames[0]]
        reqs_sheet = result_file[result_file.sheetnames[0]]

# Находим строку с заголовками
        for row in range(1, data_sheet.max_row + 1):
            if data_sheet.cell(row, data_sheet.max_column - 2).value is not None:
                header_row = row
                break

# выбираем колонки название которых начинается 'Сумма'
        sum_cols = {col: None for col in range(1, data_sheet.max_column + 1) if
                    re.match('Сумма', str(data_sheet.cell(header_row, col).value))}

# выбираем колонку с максимальной суммой
        for key in sum_cols.keys():
            sum_col = sum(int(data_sheet.cell(row, key).value) for row in range(header_row + 1, data_sheet.max_row + 1))
            sum_cols[key] = sum_col
            sum_col = 0
        ofd_income_col = max(sum_cols, key=sum_cols.get)

# выбираем первую колонку в header_row в тексте которых есть дата, время
        for col in range(1, data_sheet.max_column + 1):
            if re.search(r'Дата[и, ]+время', str(data_sheet.cell(header_row, col).value)):
                date_time_col = col
                break

# выбираем колонку в header_row в тексте которых есть Номер ФД,
# (regex содержит Номер ФД), значение присваиваем в nomer_fd_col
        for col in range(1, data_sheet.max_column + 1):
            if re.search(r'(Порядковый номер ФД)|(Номер ФД \(1040\))', str(data_sheet.cell(header_row, col).value)):
                nomer_fd_col = col
                break

# Находим последнюю строку last_row со значением в колонке nomer_fd_col и печатаем
        for row in range(data_sheet.max_row + 1, 1, -1):
            if data_sheet.cell(row, nomer_fd_col).value is not None:
                last_row = row
                break

# ЗАПОЛНЕНИЕ result_file пробегаем по всем рядам ofd_file и заполняем матрицу out_data_matrix со строками =
# [counter, дата из date_time_col, nomer_fd_col, ’Парикмахерские услуги’, ofd_income_col]
        out_data_matrix = []
        first_num = result_data_sheet.cell(result_data_sheet.max_row, 1).value
        counter_corr = first_num if first_num != 1 else 0
        counter = (last_row - header_row) + counter_corr
        operation = 'Парикмахерские услуги'
        for row in range(header_row + 1, last_row+1):
            # Определяем дату документа в колонке date_time_col офд файла
            raw_doc_date = data_sheet.cell(row, date_time_col).value
            if isinstance(raw_doc_date, str):
                doc_date = f'{raw_doc_date[6:10]}-{raw_doc_date[3:5]}-{raw_doc_date[:2]}'
            elif isinstance(raw_doc_date, datetime):
                doc_date = dt.date(raw_doc_date)
            else:
                doc_date = '0000'

            matrix_row = [counter, doc_date,
                          str(data_sheet.cell(row, nomer_fd_col).value), operation,
                          int(data_sheet.cell(row, ofd_income_col).value)]
            out_data_matrix.append(matrix_row)
            counter -= 1
        for row in reversed(out_data_matrix):
            result_data_sheet.append(row)

# Обновляем ячейки M13:O13 W30:BD30 листа result_file reqs_sheet = result_file[result_file.sheetnames[0]] согласно year
        reqs_sheet['M13'] = self.year_ent.get()[2:4]
        reqs_sheet['W30'] = f'  c 1.01.{self.year_ent.get()} по 31.12.{self.year_ent.get()}'

# Сделать однообразное выравнивание текста в ячейках result_data_sheet
        for row in range(8, result_data_sheet.max_row+1):
            for col in range(1, result_data_sheet.max_column+1):
                var = result_data_sheet.cell(row, col)
                var.alignment = Alignment(horizontal='center')

        result_file.save(result)


if __name__ == "__main__":
    app = App()
    app.mainloop()
